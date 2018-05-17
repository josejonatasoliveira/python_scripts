import openpyxl
from googlesearch import search

wb = openpyxl.load_workbook('c:/docs/Industrias/Industrias.xlsx')

sheet = wb['Plan1']

for linha in range(1,sheet.max_row + 1):
    try:
        web = 'http://google.com?#q=' + str(sheet.cell(linha,1).value + ' '+sheet.cell(linha,3).value)
        
        for url in search(web,stop=1):
            print(url)
            sheet.cell(linha,4).value = url
    except TypeError:
        continue

wb.save('c:/docs/Industrias/Industrias2.xlsx')
