import openpyxl
wb = openpyxl.load_workbook('c:/docs/localizacoes1.xlsx')
sheet = wb['Plan1']

for row in range(1,sheet.max_row + 1):
    if (len(sheet.cell(row,2).value.split(',')) == 5):
        a,b,c,d,e = sheet.cell(row,2).value.split(',')
        sheet.cell(row,4).value = a
        sheet.cell(row,5).value = b
        sheet.cell(row,6).value = c
        sheet.cell(row,7).value = d
        sheet.cell(row,8).value = e
    else:
        continue

wb.save('c:/docs/localizacoes3.xlsx')
