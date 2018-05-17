import openpyxl # Biblioteca para se trabalhar com excel
import urllib.request # Biblioteca para fazer requisições
from bs4 import BeautifulSoup # Biblioteca para extrair dados em HTML
# O codigo ir extrair os dados da tag td e class 1, 2 e 3 e gravar no arquivo excel
''' <td class="column-1">Mili</td>
    <td class="column-2">PR</td>
    <td class="column-3">BRASIL</td><td class="column-4">PAVILHÃƒO AZUL</td>
    <td class="column-5">212</td><td class="column-6">D / 6</td>'''

wb = openpyxl.load_workbook('filePathOfExcel.xlsx') # Local onde se encontra o arquivo excel
site = 'http://www.apasshow.com.br/lista-de-expositores/' # Site de exemplo para extrair as empresas que participaram do ApasShow 2018
ws = urllib.request.urlopen(site) #Abre o site

sp = BeautifulSoup(ws.read(),'html.parser') # Faz leitura do site
sheet = wb['nameOfSheet'] # Nome da planilha do arquivo excel


linha = 1
for row in sp.find_all('td',attrs={'class':'column-1'}):
    sheet.cell(linha,1).value = row.text
    linha += 1
linha = 1
for row in sp.find_all('td',attrs={'class':'column-2'}):
    sheet.cell(linha,2).value = row.text
    linha += 1
linha = 1
for row in sp.find_all('td',attrs={'class':'column-3'}):
    sheet.cell(linha,3).value = row.text
    linha += 1
wb.save('nameOfFileExcel.xlsx')
